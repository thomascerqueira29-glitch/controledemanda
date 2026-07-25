import streamlit as st
import pandas as pd
import numpy as np
import folium
from folium.plugins import MarkerCluster, HeatMap
from streamlit_folium import st_folium
import io
import zipfile
import html
import re
import requests
import time
from datetime import datetime
from openpyxl.styles import Font

from database import load_core_data

# Injeção de CSS
st.markdown("""
<style>
    .block-container { padding-top: 1.5rem !important; padding-bottom: 2rem !important; }
    .stSelectbox label, .stFileUploader label, .stRadio label, .stNumberInput label, .stMultiSelect label { font-size: 14px !important; font-weight: 600 !important; color: #1A4F7C !important; }
</style>
""", unsafe_allow_html=True)

# ==========================================
# FUNÇÕES DE LIMPEZA E FORMATAÇÃO
# ==========================================
def limpar_roteirizador():
    st.session_state.roteamento_concluido = False
    st.session_state.df_routed = pd.DataFrame()
    st.session_state.bases_records = []
    st.session_state.tipo_periodo = "Dia"
    st.session_state.colunas_exibir = []
    st.session_state.col_prioridade = "TIPO NOTA"
    st.session_state.colunas_originais = []
    st.rerun()

def normalize_cols(cols):
    new_cols = []
    for c in cols:
        c = str(c).strip().upper()
        c = re.sub(r'[ÁÀÂÃÄ]', 'A', c)
        c = re.sub(r'[ÉÈÊË]', 'E', c)
        c = re.sub(r'[ÍÌÎÏ]', 'I', c)
        c = re.sub(r'[ÓÒÔÕÖ]', 'O', c)
        c = re.sub(r'[ÚÙÛÜ]', 'U', c)
        c = re.sub(r'Ç', 'C', c)
        new_cols.append(c)
    return new_cols

def normalizar_municipios(series_mun):
    s = series_mun.astype(str).str.upper()
    s = s.str.replace(r'[ÁÀÂÃÄ]', 'A', regex=True)
    s = s.str.replace(r'[ÉÈÊË]', 'E', regex=True)
    s = s.str.replace(r'[ÍÌÎÏ]', 'I', regex=True)
    s = s.str.replace(r'[ÓÒÔÕÖ]', 'O', regex=True)
    s = s.str.replace(r'[ÚÙÛÜ]', 'U', regex=True)
    s = s.str.replace(r'Ç', 'C', regex=True)
    return s.str.split('-').str[0].str.strip()

def atualizar_status_via_arquivo(df_principal, arquivo_status):
    try:
        df_status = pd.read_excel(arquivo_status)
        if df_status.shape[1] >= 5:
            chave_nome = df_status.columns[0]
            status_nome = df_status.columns[4]
            df_status[chave_nome] = df_status[chave_nome].astype(str).str.strip()
            df_status_map = df_status.set_index(chave_nome)[status_nome].to_dict()
            if 'PROTOCOLO' in df_principal.columns:
                df_principal['PROTOCOLO_STR'] = df_principal['PROTOCOLO'].astype(str).str.strip()
                df_principal['STATUS LIST'] = df_principal['PROTOCOLO_STR'].map(df_status_map).fillna(df_principal.get('STATUS LIST', 'SEM INFORMAÇÕES'))
                df_principal = df_principal.drop(columns=['PROTOCOLO_STR'])
                st.success(f"✅ Atualização Rápida: {len(df_status_map)} status lidos da Coluna E aplicados com sucesso!")
            else:
                st.warning("⚠️ Coluna 'PROTOCOLO' não encontrada na base principal.")
        else:
            st.warning("⚠️ O arquivo de status enviado possui menos de 5 colunas.")
    except Exception as e:
        st.error(f"Erro ao ler o arquivo de atualização rápida: {e}")
    return df_principal

# ==========================================
# FUNÇÕES MATEMÁTICAS E IA (VRP / TSP 2-Opt)
# ==========================================
def haversine_vectorized(lat1, lon1, lat2, lon2):
    R = 6371.0 
    lat1, lon1, lat2, lon2 = map(np.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = np.sin(dlat/2.0)**2 + np.cos(lat1) * np.cos(lat2) * np.sin(dlon/2.0)**2
    c = 2 * np.arcsin(np.sqrt(a))
    return R * c

def kmeans_clustering(coords, k, max_iters=100):
    np.random.seed(42)
    unique_coords = np.unique(coords, axis=0)
    if len(unique_coords) < k: k = len(unique_coords)
    indices = np.random.choice(len(unique_coords), k, replace=False)
    centroids = unique_coords[indices]
    labels = np.zeros(len(coords))
    for _ in range(max_iters):
        diff = coords[:, np.newaxis, :] - centroids[np.newaxis, :, :]
        dists = np.linalg.norm(diff, axis=2)
        labels = np.argmin(dists, axis=1)
        new_centroids = np.array([coords[labels == i].mean(axis=0) if np.sum(labels == i) > 0 else centroids[i] for i in range(k)])
        if np.allclose(centroids, new_centroids): break
        centroids = new_centroids
    return labels, centroids

def otimizar_rota_tsp_2opt(lista_obras, start_lat, start_lon):
    if len(lista_obras) <= 2: return lista_obras
    coords = [(start_lat, start_lon)] + [(r['LATITUDE'], r['LONGITUDE']) for r in lista_obras]
    best_route = list(range(1, len(coords)))
    def calc_dist(route):
        d = haversine_vectorized(coords[0][0], coords[0][1], coords[route[0]][0], coords[route[0]][1])
        for i in range(len(route)-1):
            d += haversine_vectorized(coords[route[i]][0], coords[route[i]][1], coords[route[i+1]][0], coords[route[i+1]][1])
        return d
    best_dist = calc_dist(best_route)
    improved = True
    iters = 0
    while improved and iters < 50:
        improved = False
        for i in range(len(best_route) - 1):
            for j in range(i + 1, len(best_route)):
                new_route = best_route[:i] + best_route[i:j+1][::-1] + best_route[j+1:]
                new_dist = calc_dist(new_route)
                if new_dist < best_dist:
                    best_dist = new_dist
                    best_route = new_route
                    improved = True
        iters += 1
    return [lista_obras[i-1] for i in best_route]

@st.cache_data(show_spinner=False)
def obter_coordenadas_municipio_cached(municipio):
    if not municipio or pd.isna(municipio) or str(municipio).strip() == "": return np.nan, np.nan
    try:
        time.sleep(1.2)
        url = f"https://nominatim.openstreetmap.org/search?q={str(municipio).strip()},+Maranhão,+Brasil&format=json&limit=1"
        r = requests.get(url, headers={"User-Agent": "GeradorRotasOperacional/8.0"}, timeout=5)
        if r.status_code == 200 and len(r.json()) > 0: return float(r.json()[0]['lat']), float(r.json()[0]['lon'])
    except: pass
    return np.nan, np.nan

@st.cache_data(show_spinner=False)
def obter_clima_seguro(lat, lon):
    try:
        r = requests.get(f"https://wttr.in/{lat},{lon}?format=%C+%t", timeout=2)
        if r.status_code == 200: return r.text.strip()
    except: pass
    return "Dados Climáticos Indisponíveis"

def obter_rota_ruas(lat1, lon1, lat2, lon2, vel_fallback_kmh=30):
    try:
        url = f"http://router.project-osrm.org/route/v1/driving/{lon1:.6f},{lat1:.6f};{lon2:.6f},{lat2:.6f}?overview=full&geometries=geojson"
        r = requests.get(url, headers={"User-Agent": "GeradorRotasOperacional/8.0"}, timeout=5)
        if r.status_code == 200 and r.json()['code'] == 'Ok':
            return r.json()['routes'][0]['geometry']['coordinates'], r.json()['routes'][0]['duration']
    except: pass
    dist_km = haversine_vectorized(lat1, lon1, lat2, lon2)
    return [[lon1, lat1], [lon2, lat2]], (dist_km / vel_fallback_kmh) * 3600

def calcular_materiais_necessarios(tipo_nota):
    tipo = str(tipo_nota).upper()
    if 'NOVA' in tipo or 'LIGACAO' in tipo or 'UNI' in tipo or 'UNR' in tipo:
        return {'Medidor Monofásico': 1, 'Cabo Multiplexado (m)': 15, 'Conector Cunha': 2, 'Armação Secundária': 1, 'Fita Isolante': 1}
    elif 'MANUT' in tipo or 'REPARO' in tipo:
        return {'Cabo Multiplexado (m)': 5, 'Conector Cunha': 4, 'Fita Isolante': 1, 'Emenda': 2}
    elif 'INSP' in tipo or 'VISTORIA' in tipo or 'PRE ANALISE' in tipo or 'PRÉ ANÁLISE' in tipo:
        return {'Lacre de Segurança': 2, 'Lacre de Medidor': 1}
    return {'Kit Ferramentas / Miscelâneas': 1}

def identificar_icone_folium(row, colunas):
    tipo_str = str(row.get('TIPO LIGACAO', '')) + str(row.get('SERVICO', '')) + str(row.get('TIPO NOTA', ''))
    tipo_str = tipo_str.upper()
    if row.get('PROTOCOLO') == 'RETORNO_BASE': return 'home'
    if row.get('PROTOCOLO') == 'PAUSA_ALMOCO': return 'cutlery'
    if 'NOVA' in tipo_str or 'LIGACAO' in tipo_str or 'UNI' in tipo_str or 'UNR' in tipo_str: return 'bolt'
    if 'MANUT' in tipo_str or 'REPARO' in tipo_str: return 'wrench'
    if 'INSP' in tipo_str or 'VISTORIA' in tipo_str: return 'eye-open'
    return 'info-sign'

def gerar_os_html(base, periodo, df_periodo, romaneio, clima, tipo_equipe):
    linhas_tabela = ""
    for _, r in df_periodo.iterrows():
        linhas_tabela += f"<tr><td>{r.get('ORDEM','')}</td><td>{r.get('PROTOCOLO','')}</td><td>{r.get('TIPO NOTA','')}</td><td>{r.get('ENDEREÇO','')}</td></tr>"
    mat_tabela = "".join([f"<tr><td>{mat}</td><td>{qtd}</td></tr>" for mat, qtd in romaneio.items()])
    tag_temp = " (EQUIPE DE APOIO / TEMPORÁRIA)" if tipo_equipe == 'TEMPORARIA' else ""
    return f"""
    <html><head><meta charset="UTF-8"><style>
        body {{ font-family: Arial, sans-serif; padding: 20px; }}
        table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 12px; }}
        th, td {{ border: 1px solid #000; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; }}
        .header {{ border-bottom: 2px solid #000; padding-bottom: 10px; margin-bottom: 20px; }}
    </style></head><body>
        <div class="header">
            <h2>Ordem de Serviço Diária (Roteiro Otimizado)</h2>
            <p><b>Equipe:</b> {base}{tag_temp} | <b>Período:</b> {periodo} | <b>Condição Climática Prevista:</b> {clima}</p>
        </div>
        <h3>1. Roteiro de Paradas</h3>
        <table><tr><th>Ordem</th><th>Protocolo</th><th>Serviço</th><th>Endereço</th></tr>{linhas_tabela}</table>
        <h3>2. Romaneio de Carga (Almoxarifado)</h3>
        <table><tr><th>Material Necessário</th><th>Quantidade Calculada</th></tr>{mat_tabela}</table>
        <div style="display: flex; justify-content: space-around; margin-top: 60px;">
            <div style="text-align: center; border-top: 1px solid #000; width: 40%; padding-top: 5px;">Assinatura do Despachante</div>
            <div style="text-align: center; border-top: 1px solid #000; width: 40%; padding-top: 5px;">Assinatura do Técnico</div>
        </div>
    </body></html>
    """

def gerar_excel_bytes(df, col_prioridade, colunas_originais=None):
    df_export = df.copy()
    if 'ROTA_GEOMETRIA' in df_export.columns: df_export = df_export.drop(columns=['ROTA_GEOMETRIA'])
    if colunas_originais:
        cols_atuais = df_export.columns.tolist()
        cols_novas_geradas = [c for c in cols_atuais if c not in colunas_originais]
        df_export = df_export[[c for c in colunas_originais if c in cols_atuais] + cols_novas_geradas]
    buf_xl = io.BytesIO()
    with pd.ExcelWriter(buf_xl, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Roteiro')
        ws = writer.sheets['Roteiro']
        red_font = Font(color="FF0000", bold=True)
        if 'PRIORIDADE' in df_export.columns:
            prio_flag_idx = df_export.columns.get_loc('PRIORIDADE') + 1 
            for row_idx in range(2, len(df_export) + 2):
                if ws.cell(row=row_idx, column=prio_flag_idx).value == "Sim":
                    ws.cell(row=row_idx, column=prio_flag_idx).font = red_font
                    if col_prioridade != "Nenhuma" and col_prioridade in df_export.columns:
                        ws.cell(row=row_idx, column=df_export.columns.get_loc(col_prioridade) + 1).font = red_font
    return buf_xl.getvalue()

def gerar_kml_agrupado(df_rota, bases_records, doc_name, cols_exibir):
    kml = f'''<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
<Document>
  <name>{doc_name}</name>
  <Style id="linha-rota-contorno"><LineStyle><color>ff000000</color><width>8</width></LineStyle></Style>
  <Style id="linha-rota-centro"><LineStyle><color>ff00ffff</color><width>4</width></LineStyle></Style>
  <Style id="icon-blue">
    <IconStyle><scale>1.1</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/blu-blank.png</href></Icon><hotSpot x="32" xunits="pixels" y="64" yunits="insetPixels"/></IconStyle>
    <LabelStyle><scale>0.9</scale></LabelStyle>
  </Style>
  <Style id="icon-red">
    <IconStyle><scale>1.3</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-blank.png</href></Icon><hotSpot x="32" xunits="pixels" y="64" yunits="insetPixels"/></IconStyle>
    <LabelStyle><scale>1.0</scale></LabelStyle>
  </Style>
  <Style id="icon-green"><IconStyle><scale>1.2</scale><Icon><href>https://maps.google.com/mapfiles/kml/shapes/homegardenbusiness.png</href></Icon></IconStyle></Style>
  <Style id="icon-yellow"><IconStyle><scale>1.3</scale><Icon><href>https://maps.google.com/mapfiles/kml/shapes/dining.png</href></Icon><LabelStyle><scale>1.0</scale></LabelStyle></IconStyle></Style>
'''
    for base_nome in df_rota['BASE_ATRIBUIDA'].unique():
        df_base = df_rota[df_rota['BASE_ATRIBUIDA'] == base_nome]
        base_ref = next((b for b in bases_records if b['LEVANTADOR'] == base_nome), None)
        b_lat, b_lon = float(str(base_ref['LATITUDE']).replace(',','.')), float(str(base_ref['LONGITUDE']).replace(',','.'))
        res_nome = str(base_ref.get('RESIDENCIA', base_nome))

        kml += f'  <Folder>\n    <name>Levantador: {html.escape(str(base_nome))}</name>\n'
        kml += f'    <Placemark><name>BASE: {html.escape(str(res_nome))}</name><styleUrl>#icon-green</styleUrl><Point><coordinates>{b_lon},{b_lat},0</coordinates></Point></Placemark>\n'

        for semana in df_base['SEMANA'].unique():
            df_semana = df_base[df_base['SEMANA'] == semana]
            kml += f'    <Folder>\n      <name>Semana {semana}</name>\n'

            for dia in df_semana['DIA'].unique():
                df_dia = df_semana[df_semana['DIA'] == dia].copy().sort_values(by='ORDEM')
                kml += f'      <Folder>\n        <name>Dia {dia}</name>\n'

                coords_linha_kml = ""
                for _, row in df_dia.iterrows():
                    lon, lat = str(row['LONGITUDE']).replace(',','.'), str(row['LATITUDE']).replace(',','.')
                    desc_parts = [f"<b>Ordem na Rota:</b> {row.get('ORDEM', 0)}", f"<b>Distância do Ponto Anterior:</b> {row.get('DISTANCIA_PONTO_ANTERIOR_KM', 0)} KM", f"<b>Distância do Próximo Ponto:</b> {row.get('DISTANCIA_PROXIMO_PONTO_KM', 0)} KM", f"<b>Tempo de Viagem Estimado:</b> {row.get('TEMPO_VIAGEM_MINUTOS', 0)} Minutos"]
                    
                    if row.get('PROTOCOLO') == 'RETORNO_BASE':
                        desc_cdata, nome_ponto, style_url = "<b>RETORNO À BASE DE ORIGEM</b>", "🏠 FIM DO DIA - RETORNO", "#icon-green"
                    elif row.get('PROTOCOLO') == 'PAUSA_ALMOCO':
                        desc_cdata, nome_ponto, style_url = "<b>PAUSA PROGRAMADA PARA REFEIÇÃO (1h)</b>", "🍔 ALMOÇO DA EQUIPE", "#icon-yellow"
                    else:
                        for col in cols_exibir:
                            if col in row: desc_parts.append(f"<b>{col}:</b> {html.escape(str(row[col]))}")
                        desc_cdata = "<br>".join(desc_parts)
                        tag_prio = "[PRIORIDADE] " if row.get('PRIORIDADE') == "Sim" else ""
                        nome_ponto = f"{tag_prio}[{row.get('ORDEM', 0)}] Prot: {html.escape(str(row.get('PROTOCOLO', 'Sem Protocolo')))}"
                        style_url = "#icon-red" if row.get('PRIORIDADE') == "Sim" else "#icon-blue"

                    kml += f'        <Placemark><name>{nome_ponto}</name><description><![CDATA[{desc_cdata}]]></description><styleUrl>{style_url}</styleUrl><Point><coordinates>{lon},{lat},0</coordinates></Point></Placemark>\n'
                    if isinstance(row.get('ROTA_GEOMETRIA'), list):
                        coords_linha_kml += "".join([f"          {pt_lon},{pt_lat},0\n" for pt_lon, pt_lat in row['ROTA_GEOMETRIA']])
                    else:
                        coords_linha_kml += f"          {lon},{lat},0\n"

                kml += f'        <Placemark><name>Contorno Rota</name><styleUrl>#linha-rota-contorno</styleUrl><LineString><tessellate>1</tessellate><coordinates>\n{coords_linha_kml}            </coordinates></LineString></Placemark>\n' 
                kml += f'        <Placemark><name>Traçado Rota</name><styleUrl>#linha-rota-centro</styleUrl><LineString><tessellate>1</tessellate><coordinates>\n{coords_linha_kml}            </coordinates></LineString></Placemark>\n      </Folder>\n' 
            kml += '    </Folder>\n' 
        kml += '  </Folder>\n' 
    kml += '</Document>\n</kml>'
    return kml

# ==========================================
# VIEW PRINCIPAL DA PÁGINA
# ==========================================
def view_roteirizador():
    if "roteamento_concluido" not in st.session_state:
        st.session_state.roteamento_concluido = False
    if "df_routed" not in st.session_state:
        st.session_state.df_routed = pd.DataFrame()
    if "bases_records" not in st.session_state:
        st.session_state.bases_records = []
    if "tipo_periodo" not in st.session_state:
        st.session_state.tipo_periodo = "Dia"
    if "colunas_exibir" not in st.session_state:
        st.session_state.colunas_exibir = []
    if "col_prioridade" not in st.session_state:
        st.session_state.col_prioridade = "TIPO NOTA"
    if "colunas_originais" not in st.session_state:
        st.session_state.colunas_originais = []

    if st.session_state.roteamento_concluido and not st.session_state.df_routed.empty:
        st.markdown("## 🎯 Resultados da Roteirização Corporativa")
        st.markdown("### ✍️ Ajuste Fino Manual (Painel do Despachante)")
        st.info("Dê um **duplo clique** nas células abaixo para alterar o responsável ou a ordem das obras. Suas edições sairão direto nos downloads finais.")
        
        df_editado_ui = st.data_editor(
            st.session_state.df_routed, use_container_width=True,
            column_config={ "ROTA_GEOMETRIA": None, "LATITUDE": st.column_config.NumberColumn(disabled=True), "LONGITUDE": st.column_config.NumberColumn(disabled=True), "DISTANCIA_PONTO_ANTERIOR_KM": st.column_config.NumberColumn(disabled=True), "DISTANCIA_PROXIMO_PONTO_KM": st.column_config.NumberColumn(disabled=True), "TEMPO_VIAGEM_MINUTOS": st.column_config.NumberColumn(disabled=True) }
        )
        
        df_routed = df_editado_ui.copy()
        bases_records = st.session_state.bases_records
        tipo_periodo = st.session_state.tipo_periodo
        colunas_exibir = st.session_state.colunas_exibir
        col_prioridade = st.session_state.col_prioridade
        colunas_originais = st.session_state.colunas_originais
        
        df_real_tasks = df_routed[~df_routed['PROTOCOLO'].isin(['RETORNO_BASE', 'PAUSA_ALMOCO'])]
        
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("📌 Obras Roteirizadas", len(df_real_tasks))
        k2.metric("👥 Equipes em Campo", df_routed['BASE_ATRIBUIDA'].nunique())
        k3.metric("🛣️ KM Total Projetado", f"{df_routed['DISTANCIA_PONTO_ANTERIOR_KM'].sum():.1f} km")
        k4.metric("🚨 Prioridades", len(df_real_tasks[df_real_tasks['PRIORIDADE'] == 'Sim']) if 'PRIORIDADE' in df_real_tasks else 0)

        st.markdown("---")
        st.markdown("### 📊 Dashboards de Produtividade")
        c_dash1, c_dash2 = st.columns(2)
        with c_dash1:
            st.markdown("##### 📦 Volume de Obras por Equipe")
            st.bar_chart(df_real_tasks['BASE_ATRIBUIDA'].value_counts(), color="#1A4F7C")
        with c_dash2:
            st.markdown("##### 🛣️ Quilometragem Projetada por Equipe")
            st.bar_chart(df_routed.groupby('BASE_ATRIBUIDA')['DISTANCIA_PONTO_ANTERIOR_KM'].sum(), color="#FF4B4B")
        st.markdown("---")

        st.markdown("#### 🗺️ Visualização Geográfica do Plano")
        mapa = folium.Map(location=[df_routed['LATITUDE'].mean(), df_routed['LONGITUDE'].mean()], zoom_start=8) if not df_routed.empty else folium.Map(location=[-5.2, -45.0], zoom_start=7)
        cores = ['#f1c40f', '#00b894', '#9b59b6', '#e67e22', '#e74c3c', '#1abc9c', '#27ae60', '#2980b9']
        
        heat_data = [[r['LATITUDE'], r['LONGITUDE']] for _, r in df_real_tasks.iterrows()]
        HeatMap(heat_data, name="🔥 Mapa de Calor (Demandas)", radius=15, blur=10).add_to(mapa)
        
        marker_cluster = MarkerCluster(name="Obras (Agrupadas)").add_to(mapa)
        
        for idx, base_nome in enumerate(df_routed['BASE_ATRIBUIDA'].unique()):
            cor_rota = cores[idx % len(cores)]
            df_base_rota = df_routed[df_routed['BASE_ATRIBUIDA'] == base_nome]
            base_ref = next((b for b in bases_records if b['LEVANTADOR'] == base_nome), None)
            b_lat, b_lon = float(str(base_ref['LATITUDE']).replace(',','.')), float(str(base_ref['LONGITUDE']).replace(',','.'))
            folium.Marker([b_lat, b_lon], icon=folium.Icon(color='black', icon='home', prefix='fa'), tooltip=f"Base: {base_nome}").add_to(mapa)
            
            for periodo_val in df_base_rota['PERIODO'].unique():
                df_periodo = df_base_rota[df_base_rota['PERIODO'] == periodo_val]
                fg_linhas = folium.FeatureGroup(name=f"Linhas {base_nome} | P: {periodo_val}", show=False)
                
                pontos_linha_folium = []
                for _, r in df_periodo.iterrows():
                    if isinstance(r.get('ROTA_GEOMETRIA'), list):
                        for lon, lat in r['ROTA_GEOMETRIA']: pontos_linha_folium.append([lat, lon]) 
                            
                folium.PolyLine(pontos_linha_folium, color='black', weight=7, opacity=0.9).add_to(fg_linhas)
                folium.PolyLine(pontos_linha_folium, color=cor_rota, weight=3, opacity=1.0).add_to(fg_linhas)
                fg_linhas.add_to(mapa)
                
                for _, r in df_periodo.iterrows():
                    if r['PROTOCOLO'] in ['RETORNO_BASE', 'PAUSA_ALMOCO']: continue
                    icone = identificar_icone_folium(r, df_routed.columns)
                    cor_icone = 'red' if r.get('PRIORIDADE') == "Sim" else 'blue'
                    
                    info_html = f"<b>Ordem:</b> {r.get('ORDEM', 0)} | <b>{tipo_periodo}:</b> {r.get('PERIODO', 0)}<br><b>Distância Próximo Ponto:</b> {r.get('DISTANCIA_PROXIMO_PONTO_KM', 0)} KM<br><b>Tempo Estimado:</b> {r.get('TEMPO_VIAGEM_MINUTOS', 0)} Min<br>"
                    for c in colunas_exibir:
                        if c in r: info_html += f"<b>{c}:</b> {r[c]}<br>"
                        
                    folium.Marker([r['LATITUDE'], r['LONGITUDE']], icon=folium.Icon(color=cor_icone, icon=icone), popup=folium.Popup(info_html, max_width=300)).add_to(marker_cluster)
        
        folium.LayerControl().add_to(mapa)
        st_folium(mapa, use_container_width=True, height=550, returned_objects=[])

        st.markdown("#### 📥 Baixar Resultados e Integrações")
        data_atual = datetime.now().strftime("%d_%m_%Y")
        
        buf_zip_xl = io.BytesIO()
        with zipfile.ZipFile(buf_zip_xl, 'w', zipfile.ZIP_DEFLATED) as zip_xl:
            # 1. Roteiro Geral & PowerBI
            zip_xl.writestr(f"Roteiro_Geral_{data_atual}.xlsx", gerar_excel_bytes(df_routed, col_prioridade, colunas_originais))
            planilhas_geradas = [f"Roteiro_Geral_{data_atual}.xlsx"]
            
            cols_atuais_bi = df_routed.columns.tolist()
            cols_novas_bi = [c for c in cols_atuais_bi if c not in colunas_originais]
            zip_xl.writestr(f"Base_Dashboards_PowerBI_{data_atual}.csv", df_routed[[c for c in colunas_originais if c in cols_atuais_bi] + cols_novas_bi].to_csv(index=False, sep=';', decimal=',', encoding='utf-8-sig'))
            planilhas_geradas.append(f"Base_Dashboards_PowerBI_{data_atual}.csv")
            
            # 2. Layout SAP
            sap_cols = [c for c in ['PROTOCOLO', 'ORDEM', 'BASE_ATRIBUIDA', 'TIPO LIGACAO', 'STATUS SAP'] if c in df_real_tasks.columns]
            if sap_cols:
                df_sap = df_real_tasks[sap_cols].copy()
                df_sap['NOVO_STATUS_ATUALIZACAO'] = ''
                zip_xl.writestr(f"Layout_Importacao_SAP_{data_atual}.xlsx", gerar_excel_bytes(df_sap, "Nenhuma"))
                planilhas_geradas.append(f"Layout_Importacao_SAP_{data_atual}.xlsx")

            # 3. Expectativa, Romaneio e OS em HTML (PDF)
            resumo_data, romaneio_data = [], []
            
            for base in df_routed['BASE_ATRIBUIDA'].unique():
                df_base = df_routed[df_routed['BASE_ATRIBUIDA'] == base]
                base_ref = next((b for b in bases_records if b['LEVANTADOR'] == base), None)
                clima_base = obter_clima_seguro(df_base.iloc[0]['LATITUDE'], df_base.iloc[0]['LONGITUDE'])
                
                for periodo in df_base['PERIODO'].unique():
                    df_periodo = df_base[df_base['PERIODO'] == periodo]
                    df_periodo_real = df_periodo[~df_periodo['PROTOCOLO'].isin(['RETORNO_BASE', 'PAUSA_ALMOCO'])]
                    
                    qtd_obras = len(df_periodo_real)
                    qtd_prio = len(df_periodo_real[df_periodo_real['PRIORIDADE'] == 'Sim']) if 'PRIORIDADE' in df_periodo_real.columns else 0
                    
                    resumo_data.append({
                        'LEVANTADOR': base, 'TIPO EQUIPE': base_ref.get('TIPO_EQUIPE', 'PRINCIPAL'), 
                        f'{tipo_periodo.upper()}': periodo, 'OBRAS ROTEIRIZADAS': qtd_obras,
                        'OBRAS PRIORITARIAS': qtd_prio, 'KM TOTAL PROJETADO': round(df_periodo['DISTANCIA_PONTO_ANTERIOR_KM'].sum(), 2),
                        'LINK GPS (GOOGLE MAPS)': "https://www.google.com/maps/dir/" + "/".join([f"{lat},{lon}" for lat, lon in df_periodo[['LATITUDE', 'LONGITUDE']].values.tolist()])
                    })
                    
                    romaneio_periodo = {}
                    for _, r in df_periodo_real.iterrows():
                        for mat, qtd in calcular_materiais_necessarios(r.get('TIPO NOTA', 'Misto')).items():
                            romaneio_periodo[mat] = romaneio_periodo.get(mat, 0) + qtd
                            romaneio_data.append({'LEVANTADOR / EQUIPE': base, f'{tipo_periodo.upper()}': periodo, 'MATERIAL': mat, 'QUANTIDADE': qtd})
                            
                    os_html = gerar_os_html(base, f"{tipo_periodo} {periodo}", df_periodo_real, romaneio_periodo, clima_base, base_ref.get('TIPO_EQUIPE', 'PRINCIPAL'))
                    nome_seg_os = re.sub(r'[^A-Za-z0-9_]', '', str(base).replace(" ", "_"))
                    zip_xl.writestr(f"Ordens_Servico_Imprimir/OS_{nome_seg_os}_{tipo_periodo}{periodo}.html", os_html.encode('utf-8'))

            buf_resumo = io.BytesIO()
            with pd.ExcelWriter(buf_resumo, engine='openpyxl') as writer:
                pd.DataFrame(resumo_data).to_excel(writer, index=False, sheet_name='Resumo')
            zip_xl.writestr(f"Expectativa_{'Semanal' if tipo_periodo == 'Semana' else 'Diaria'}_{data_atual}.xlsx", buf_resumo.getvalue())
            planilhas_geradas.append(f"Expectativa_{'Semanal' if tipo_periodo == 'Semana' else 'Diaria'}_{data_atual}.xlsx")
            
            if romaneio_data:
                buf_romaneio = io.BytesIO()
                with pd.ExcelWriter(buf_romaneio, engine='openpyxl') as writer:
                    pd.DataFrame(romaneio_data).groupby(['LEVANTADOR / EQUIPE', f'{tipo_periodo.upper()}', 'MATERIAL'])['QUANTIDADE'].sum().reset_index().to_excel(writer, index=False, sheet_name='Almoxarifado')
                zip_xl.writestr(f"Romaneio_Materiais_{data_atual}.xlsx", buf_romaneio.getvalue())
                planilhas_geradas.append(f"Romaneio_Materiais_{data_atual}.xlsx")

            # 4. Planilhas Individuais
            for base_nome in df_routed['BASE_ATRIBUIDA'].unique():
                df_lev = df_routed[df_routed['BASE_ATRIBUIDA'] == base_nome].copy()
                nome_seguro = re.sub(r'[^A-Za-z0-9_]', '', str(base_nome).replace(" ", "_"))
                if not df_lev.empty:
                    zip_xl.writestr(f"Roteiro_{nome_seguro}_{data_atual}.xlsx", gerar_excel_bytes(df_lev, col_prioridade, colunas_originais))
                    planilhas_geradas.append(f"Roteiro_{nome_seguro}_{data_atual}.xlsx")
                    
        zip_xl_bytes = buf_zip_xl.getvalue()

        buf_zip_kml = io.BytesIO()
        with zipfile.ZipFile(buf_zip_kml, 'w', zipfile.ZIP_DEFLATED) as zip_kml:
            zip_kml.writestr(f"Rota_Geral_{data_atual}.kml", gerar_kml_agrupado(df_routed, bases_records, f"Rota_Geral_{data_atual}", colunas_exibir).encode('utf-8'))
            mapas_gerados = [f"Rota_Geral_{data_atual}.kml"]
            
            for base_nome in df_routed['BASE_ATRIBUIDA'].unique():
                df_lev = df_routed[df_routed['BASE_ATRIBUIDA'] == base_nome].copy()
                nome_seguro = re.sub(r'[^A-Za-z0-9_]', '', str(base_nome).replace(" ", "_"))
                if not df_lev.empty:
                    zip_kml.writestr(f"Rota_{nome_seguro}_{data_atual}.kml", gerar_kml_agrupado(df_lev, bases_records, f"Rota_{nome_seguro}", colunas_exibir).encode('utf-8'))
                    mapas_gerados.append(f"Rota_{nome_seguro}_{data_atual}.kml")
        zip_kml_bytes = buf_zip_kml.getvalue()

        with st.expander("📄 Ver lista de arquivos gerados (Conteúdo dos ZIPs)"):
            st.markdown("**Planilhas Excel:** " + ", ".join(planilhas_geradas))
            st.markdown("**Mapas KML:** " + ", ".join(mapas_gerados))
            st.markdown("**Ordens de Serviço (Para Impressão):** Geradas em HTML dentro do ZIP.")

        col_b1, col_b2, col_b3 = st.columns([1, 1, 1])
        col_b1.download_button("🌐 1. Planilhas, BI, OS e Romaneio (ZIP)", data=zip_xl_bytes, file_name=f"Dados_Estruturados_Roteiro_{data_atual}.zip", mime="application/zip", use_container_width=True)
        col_b2.download_button("🗺️ 2. Baixar Mapas (KML ZIP)", data=zip_kml_bytes, file_name=f"Mapas_KML_{data_atual}.zip", mime="application/zip", use_container_width=True)
        if col_b3.button("🧹 Zerar Roteirizador", type="primary", use_container_width=True):
            limpar_roteirizador()
        
        return 

    # -------------------------------------------------------------
    # TELA DE CONFIGURAÇÃO INICIAL
    # -------------------------------------------------------------
    st.markdown("## 🚙 Roteirizador Operacional Avançado")
    st.markdown("Planeje rotas inteligentes integradas a controles de esforço e retorno à base.")

    with st.sidebar:
        st.markdown("### ⚙️ Gestão de Esforço Diário")
        tipo_periodo = st.radio("Como agrupar o roteiro?", ["Dia", "Semana"], horizontal=True)
        modo_limite = st.radio("Critério limitador da equipe:", ["Quantidade Fixa de Obras", "Carga Horária (Tempo Real via Satélite)"])
        
        obras_por_periodo = 10
        horas_por_dia = 8.0
        tempo_medio_obra = 1.5
        velocidade_media_kmh = 30.0
        
        if modo_limite == "Quantidade Fixa de Obras":
            obras_por_periodo = st.number_input(f"Máximo de Obras por {tipo_periodo}", min_value=1, value=10, step=1)
            limite_periodos = st.number_input(f"Limite total de {tipo_periodo}s a roteirizar", min_value=1, value=5, step=1)
        else:
            horas_por_dia = st.number_input(f"Horas de trabalho disponíveis por {tipo_periodo}", min_value=1.0, value=8.0, step=0.5)
            tempo_medio_obra = st.number_input("Tempo médio de execução por obra (Horas)", min_value=0.1, value=1.5, step=0.1)
            velocidade_media_kmh = st.number_input("Velocidade (Plano B de Conexão) (km/h)", min_value=10.0, value=30.0, step=5.0)
            limite_periodos = st.number_input(f"Limite total de {tipo_periodo}s a roteirizar", min_value=1, value=5, step=1)

    col_up_1, col_up_2 = st.columns(2)

    with col_up_1:
        st.markdown("### 👥 1. Gestão de Equipes (Bases)")
        origem_bases = st.radio("Fonte dos Levantadores", ["Banco de Dados do Sistema", "Upload Planilha Levantadores_MA"])
        df_bases = pd.DataFrame()

        if origem_bases == "Banco de Dados do Sistema":
            _, df_equipes_db, _, _, _, _, _, _ = load_core_data()
            if not df_equipes_db.empty:
                df_equipes_db.columns = normalize_cols(df_equipes_db.columns)
                if 'LEVANTADOR' not in df_equipes_db.columns:
                    for p_nome in ['NOME', 'TECNICO', 'EQUIPE', 'COLABORADOR']:
                        if p_nome in df_equipes_db.columns:
                            df_equipes_db = df_equipes_db.rename(columns={p_nome: 'LEVANTADOR'})
                            break
                            
                if 'LEVANTADOR' in df_equipes_db.columns:
                    if 'RESIDENCIA' in df_equipes_db.columns:
                        muns_unicos = df_equipes_db['RESIDENCIA'].dropna().unique()
                        mapa_coords = {}
                        with st.spinner("🌍 Mapeando coordenadas dos municípios-base (Satélite)..."):
                            for mun in muns_unicos:
                                if str(mun).strip() != "":
                                    lat, lon = obter_coordenadas_municipio_cached(mun)
                                    mapa_coords[mun] = (lat, lon)
                        df_equipes_db['LATITUDE'] = df_equipes_db['RESIDENCIA'].map(lambda x: mapa_coords.get(x, (np.nan, np.nan))[0])
                        df_equipes_db['LONGITUDE'] = df_equipes_db['RESIDENCIA'].map(lambda x: mapa_coords.get(x, (np.nan, np.nan))[1])
                    else:
                        df_equipes_db['LATITUDE'] = pd.to_numeric(df_equipes_db.get('LATITUDE', pd.Series()).astype(str).str.replace(',', '.'), errors='coerce')
                        df_equipes_db['LONGITUDE'] = pd.to_numeric(df_equipes_db.get('LONGITUDE', pd.Series()).astype(str).str.replace(',', '.'), errors='coerce')

                    lista_lev = sorted([str(x) for x in df_equipes_db['LEVANTADOR'].dropna().unique().tolist()])
                    levs_selecionados = st.multiselect("Selecione as Equipes que irão a campo:", lista_lev)
                    if levs_selecionados:
                        df_bases = df_equipes_db[df_equipes_db['LEVANTADOR'].isin(levs_selecionados)].copy()
                        df_bases = df_bases.dropna(subset=['LATITUDE', 'LONGITUDE'])
                        df_bases['TIPO_EQUIPE'] = 'PRINCIPAL'
                        if len(df_bases) < len(levs_selecionados):
                            st.warning("⚠️ Alguns levantadores principais foram ignorados pois o município não foi localizado.")
                else:
                    st.error("❌ A coluna 'LEVANTADOR' não foi encontrada no Banco de Dados.")
        else:
            base_file = st.file_uploader("Suba a planilha Levantadores_MA", type=["xlsx", "xls"])
            if base_file:
                try:
                    df_bases_temp_ui = pd.read_excel(base_file)
                    df_bases_temp_ui.columns = normalize_cols(df_bases_temp_ui.columns)
                    if 'LEVANTADOR' not in df_bases_temp_ui.columns:
                        for p_nome in ['NOME', 'TECNICO', 'EQUIPE', 'COLABORADOR']:
                            if p_nome in df_bases_temp_ui.columns:
                                df_bases_temp_ui = df_bases_temp_ui.rename(columns={p_nome: 'LEVANTADOR'})
                                break
                    if 'LEVANTADOR' in df_bases_temp_ui.columns:
                        opcoes_levs = sorted([str(x) for x in df_bases_temp_ui['LEVANTADOR'].dropna().unique().tolist() if str(x).upper().strip() != 'SEM LEVANTADOR'])
                        levs_selecionados = st.multiselect("Selecione as Equipes Principais:", opcoes_levs)
                        if levs_selecionados:
                            df_bases = df_bases_temp_ui[df_bases_temp_ui['LEVANTADOR'].isin(levs_selecionados)].copy()
                            if 'RESIDENCIA' in df_bases.columns:
                                muns_unicos = df_bases['RESIDENCIA'].dropna().unique()
                                mapa_coords = {}
                                with st.spinner("🌍 Mapeando coordenadas dos municípios-base (Satélite)..."):
                                    for mun in muns_unicos:
                                        lat, lon = obter_coordenadas_municipio_cached(mun)
                                        mapa_coords[mun] = (lat, lon)
                                df_bases['LATITUDE'] = df_bases['RESIDENCIA'].map(lambda x: mapa_coords.get(x, (np.nan, np.nan))[0])
                                df_bases['LONGITUDE'] = df_bases['RESIDENCIA'].map(lambda x: mapa_coords.get(x, (np.nan, np.nan))[1])
                            else:
                                df_bases['LATITUDE'] = pd.to_numeric(df_bases.get('LATITUDE', pd.Series()).astype(str).str.replace(',', '.'), errors='coerce')
                                df_bases['LONGITUDE'] = pd.to_numeric(df_bases.get('LONGITUDE', pd.Series()).astype(str).str.replace(',', '.'), errors='coerce')
                                
                            df_bases = df_bases.dropna(subset=['LATITUDE', 'LONGITUDE'])
                            df_bases['TIPO_EQUIPE'] = 'PRINCIPAL'
                except Exception as e:
                    st.error(f"Erro ao ler a planilha: {e}")

        st.markdown("##### Regra de Atribuição Territorial")
        tipo_atribuicao = st.radio("Regra", ["Clusterização Inteligente por IA (K-Means VRP)", "Por Proximidade Geográfica das Coordenadas (Ignora texto)", "Por Municípios Atendidos (Lê texto da planilha)"], label_visibility="collapsed")

    with col_up_2:
        st.markdown("### 📁 2. Upload de Demandas (Obras)")
        task_files = st.file_uploader("1️⃣ Base Principal (Planilha de Obras Antiga/Original)", type=["xlsx", "xls"], accept_multiple_files=True)
        
        st.markdown("##### 🔄 Atualização Rápida de Status (Opcional)")
        status_file = st.file_uploader("2️⃣ Planilha Atualizada do SharePoint (Atualiza a Coluna E)", type=["xlsx", "xls"])
        
        # --- NOVO: BLOCO DOS TEMPORÁRIOS ---
        st.markdown("##### 🧑‍🤝‍🧑 3. Equipes de Apoio (Temporários - Opcional)")
        st.caption("Recebem APENAS obras comuns. O volume de trabalho é dividido nas mesmas regiões das equipes principais.")
        temp_bases_files = st.file_uploader("Suba a(s) planilha(s) de Levantadores Temporários", type=["xlsx", "xls"], accept_multiple_files=True)
        
        df_bases_temp = pd.DataFrame()
        if temp_bases_files:
            try:
                dfs_temp = []
                for f in temp_bases_files:
                    df_t = pd.read_excel(f)
                    df_t.columns = normalize_cols(df_t.columns)
                    if 'LEVANTADOR' not in df_t.columns:
                        for p_nome in ['NOME', 'TECNICO', 'EQUIPE', 'COLABORADOR']:
                            if p_nome in df_t.columns:
                                df_t = df_t.rename(columns={p_nome: 'LEVANTADOR'})
                                break
                    dfs_temp.append(df_t)
                df_bases_temp_full = pd.concat(dfs_temp, ignore_index=True)
                
                if 'LEVANTADOR' in df_bases_temp_full.columns:
                    opcoes_levs_temp = sorted([str(x) for x in df_bases_temp_full['LEVANTADOR'].dropna().unique().tolist() if str(x).upper().strip() != 'SEM LEVANTADOR'])
                    levs_temp_selecionados = st.multiselect("Selecione as Equipes Temporárias:", opcoes_levs_temp, key="ms_temp")
                    
                    if levs_temp_selecionados:
                        df_bases_temp = df_bases_temp_full[df_bases_temp_full['LEVANTADOR'].isin(levs_temp_selecionados)].copy()
                        if 'RESIDENCIA' in df_bases_temp.columns:
                            muns_unicos_temp = df_bases_temp['RESIDENCIA'].dropna().unique()
                            mapa_coords_temp = {}
                            with st.spinner("🌍 Mapeando bases dos temporários..."):
                                for mun in muns_unicos_temp:
                                    lat, lon = obter_coordenadas_municipio_cached(mun)
                                    mapa_coords_temp[mun] = (lat, lon)
                            df_bases_temp['LATITUDE'] = df_bases_temp['RESIDENCIA'].map(lambda x: mapa_coords_temp.get(x, (np.nan, np.nan))[0])
                            df_bases_temp['LONGITUDE'] = df_bases_temp['RESIDENCIA'].map(lambda x: mapa_coords_temp.get(x, (np.nan, np.nan))[1])
                        else:
                            df_bases_temp['LATITUDE'] = pd.to_numeric(df_bases_temp.get('LATITUDE', pd.Series()).astype(str).str.replace(',', '.'), errors='coerce')
                            df_bases_temp['LONGITUDE'] = pd.to_numeric(df_bases_temp.get('LONGITUDE', pd.Series()).astype(str).str.replace(',', '.'), errors='coerce')
                        
                        df_bases_temp = df_bases_temp.dropna(subset=['LATITUDE', 'LONGITUDE'])
                        df_bases_temp['TIPO_EQUIPE'] = 'TEMPORARIA'
            except Exception as e:
                st.error(f"Erro ao ler temporários: {e}")

        if not task_files: st.info("Aguardando upload para habilitar a configuração."); return

        try:
            dfs = []
            for f in task_files:
                df_temp = pd.read_excel(f)
                if len(dfs) == 0: st.session_state.colunas_originais = df_temp.columns.tolist()
                df_temp.columns = normalize_cols(df_temp.columns)
                dfs.append(df_temp)
            df_tasks = pd.concat(dfs, ignore_index=True)
        except Exception as e:
            st.error(f"Erro ao unificar as planilhas: {e}"); return

        if status_file:
            df_tasks = atualizar_status_via_arquivo(df_tasks, status_file)

    if 'LATITUDE' not in df_tasks.columns or 'LONGITUDE' not in df_tasks.columns:
        st.error("❌ A planilha de Obras precisa ter LATITUDE e LONGITUDE."); return

    st.markdown("---")
    
    # === LIMPEZA E MARCAÇÃO DE PRIORIDADES ===
    total_orig = len(df_tasks)
    df_tasks['LATITUDE'] = pd.to_numeric(df_tasks['LATITUDE'].astype(str).str.replace(',', '.'), errors='coerce')
    df_tasks['LONGITUDE'] = pd.to_numeric(df_tasks['LONGITUDE'].astype(str).str.replace(',', '.'), errors='coerce')
    df_tasks = df_tasks.dropna(subset=['LATITUDE', 'LONGITUDE'])
    df_tasks = df_tasks[(df_tasks['LATITUDE'] != 0.0) & (df_tasks['LONGITUDE'] != 0.0)]
    
    if 'NOME DO SOLICITANTE' in df_tasks.columns:
        df_tasks = df_tasks.dropna(subset=['NOME DO SOLICITANTE'])
        df_tasks = df_tasks[df_tasks['NOME DO SOLICITANTE'].astype(str).str.strip() != '']
    if 'STATUS SAP' in df_tasks.columns:
        df_tasks = df_tasks[~df_tasks['STATUS SAP'].astype(str).str.strip().str.upper().isin(['CANC', 'FINL'])]

    if 'STATUS LIST' in df_tasks.columns:
        status_validos = ['EM LEVANTAMENTO', '0', 'SEM INFORMAÇÕES', 'SEM INFORMACOES', 'CORREÇÃO DE LEVANTAMENTO', 'CORRECAO DE LEVANTAMENTO', 'PRÉ ANÁLISE', 'PRE ANALISE']
        df_tasks = df_tasks[df_tasks['STATUS LIST'].astype(str).str.strip().str.upper().isin(status_validos)]

    # Marca as prioridades LOGO DEPOIS da limpeza para usar na separação territorial
    tipos_prioritarios = ["CCF", "DIF", "MGD", "MTP", "ASC", "SID"]
    if 'TIPO NOTA' in df_tasks.columns:
        df_tasks['PRIORIDADE'] = df_tasks['TIPO NOTA'].apply(lambda x: 'Sim' if str(x).strip().upper() in tipos_prioritarios else 'Não')
    else:
        df_tasks['PRIORIDADE'] = 'Não'

    if total_orig - len(df_tasks) > 0:
        st.warning(f"⚠️ {total_orig - len(df_tasks)} obras com erros sistêmicos ou de Status foram ignoradas. Restam **{len(df_tasks)} válidas.**")

    if df_tasks.empty: return

    # === PRÉ-ALOCAÇÃO TERRITORIAL EXCLUSIVA ===
    df_tasks_alocadas = pd.DataFrame()
    bases_principais_records = df_bases.to_dict('records') if not df_bases.empty else []
    bases_temporarias_records = df_bases_temp.to_dict('records') if not df_bases_temp.empty else []
    todas_bases_records = bases_principais_records + bases_temporarias_records
    
    if len(todas_bases_records) > 0:
        df_tasks['BASE_ATRIBUIDA'] = "NÃO ALOCADO"
        
        # Separa a carga de trabalho: Prioridades (apenas Principais) x Comuns (Todas as Equipes)
        df_prio = df_tasks[df_tasks['PRIORIDADE'] == 'Sim'].copy()
        df_comum = df_tasks[df_tasks['PRIORIDADE'] == 'Não'].copy()
        
        if tipo_atribuicao == "Clusterização Inteligente por IA (K-Means VRP)":
            def allocate_kmeans(df_subset, base_list):
                if df_subset.empty or not base_list: return df_subset
                k = len(base_list)
                if k > 0 and len(df_subset) >= k:
                    coords = df_subset[['LATITUDE', 'LONGITUDE']].values
                    labels, centroids = kmeans_clustering(coords, k)
                    base_coords = {b['LEVANTADOR']: (float(b['LATITUDE']), float(b['LONGITUDE'])) for b in base_list if pd.notna(b.get('LATITUDE'))}
                    used_bases = set()
                    
                    for i, centroid in enumerate(centroids):
                        best_base = None
                        min_dist = float('inf')
                        for b_name, (b_lat, b_lon) in base_coords.items():
                            if b_name in used_bases: continue
                            dist = haversine_vectorized(centroid[0], centroid[1], b_lat, b_lon)
                            if dist < min_dist: min_dist, best_base = dist, b_name
                        if best_base:
                            used_bases.add(best_base)
                            df_subset.loc[df_subset.index[labels == i], 'BASE_ATRIBUIDA'] = best_base
                else:
                    # Se houver menos obras que equipes, roda por proximidade para não quebrar a IA
                    for idx, row in df_subset.iterrows():
                        best_dist, best_b = float('inf'), "NÃO ALOCADO"
                        for b in base_list:
                            d = haversine_vectorized(row['LATITUDE'], row['LONGITUDE'], float(b['LATITUDE']), float(b['LONGITUDE']))
                            if d < best_dist: best_dist, best_b = d, b['LEVANTADOR']
                        df_subset.loc[idx, 'BASE_ATRIBUIDA'] = best_b
                return df_subset

            # Distribuição segregada
            df_prio = allocate_kmeans(df_prio, bases_principais_records)
            df_comum = allocate_kmeans(df_comum, todas_bases_records)
            df_tasks = pd.concat([df_prio, df_comum])

        elif tipo_atribuicao == "Por Proximidade Geográfica das Coordenadas (Ignora texto)":
            def get_nearest_base(lat, lon, base_list):
                if not base_list: return "NÃO ALOCADO"
                min_dist, best_base = float('inf'), None
                for b in base_list:
                    if pd.notna(b.get('LATITUDE')):
                        d = haversine_vectorized(lat, lon, float(b['LATITUDE']), float(b['LONGITUDE']))
                        if d < min_dist: min_dist, best_base = d, b['LEVANTADOR']
                return best_base if best_base else "NÃO ALOCADO"
                
            df_prio['BASE_ATRIBUIDA'] = df_prio.apply(lambda r: get_nearest_base(r['LATITUDE'], r['LONGITUDE'], bases_principais_records), axis=1)
            df_comum['BASE_ATRIBUIDA'] = df_comum.apply(lambda r: get_nearest_base(r['LATITUDE'], r['LONGITUDE'], todas_bases_records), axis=1)
            df_tasks = pd.concat([df_prio, df_comum])

        elif tipo_atribuicao == "Por Municípios Atendidos (Lê texto da planilha)":
            mun_to_main = {}
            mun_to_all = {}
            
            for b in todas_bases_records:
                for m in str(b.get('MUNICIPIO', '')).split(','):
                    m_limpo = normalizar_municipios(pd.Series([m])).iloc[0]
                    if m_limpo:
                        if m_limpo not in mun_to_all: mun_to_all[m_limpo] = []
                        mun_to_all[m_limpo].append(b['LEVANTADOR'])
                        if b.get('TIPO_EQUIPE') == 'PRINCIPAL':
                            if m_limpo not in mun_to_main: mun_to_main[m_limpo] = []
                            mun_to_main[m_limpo].append(b['LEVANTADOR'])
            
            df_tasks['MUN_LIMPO'] = normalizar_municipios(df_tasks['MUNICIPIO'])
            
            def allocate_by_mun_divided(df_sub, map_dict):
                df_sub = df_sub.copy()
                df_sub['BASE_ATRIBUIDA'] = "NÃO ALOCADO"
                for mun, group in df_sub.groupby('MUN_LIMPO'):
                    bases_disp = map_dict.get(mun, [])
                    if bases_disp:
                        n_bases = len(bases_disp)
                        # Divisão no formato Round-Robin para garantir lotes justos e sem repetição
                        assigned = [bases_disp[i % n_bases] for i in range(len(group))]
                        df_sub.loc[group.index, 'BASE_ATRIBUIDA'] = assigned
                return df_sub
                
            df_prio = allocate_by_mun_divided(df_prio, mun_to_main)
            df_comum = allocate_by_mun_divided(df_comum, mun_to_all)
            
            df_tasks = pd.concat([df_prio, df_comum]).drop(columns=['MUN_LIMPO'])

        df_unallocated = df_tasks[df_tasks['BASE_ATRIBUIDA'] == "NÃO ALOCADO"]
        df_tasks_alocadas = df_tasks[df_tasks['BASE_ATRIBUIDA'] != "NÃO ALOCADO"].copy()

        if df_tasks_alocadas.empty:
            st.error("Falha: Nenhuma obra encontrada no território das equipes selecionadas. Troque a regra ou o Levantador.")
            return

        if not df_unallocated.empty:
            st.warning(f"⚠️ {len(df_unallocated)} obras carregadas ficaram sem Levantador. Motivos possíveis: Não pertencem à região das equipes ou são prioritárias e não havia equipe Principal alocada.")
            
        bases_records = todas_bases_records # Atualiza a variável master para o Engine.

    # === CONFIGURAÇÃO DE EXIBIÇÃO ===
    if not df_tasks_alocadas.empty:
        with st.expander("🛠️ 4. Configuração de Roteirização (Filtros)", expanded=True):
            c_ex1, c_ex2 = st.columns(2)
            
            if 'TIPO NOTA' in df_tasks_alocadas.columns:
                tipos_nota_unicos = sorted(df_tasks_alocadas['TIPO NOTA'].astype(str).dropna().unique().tolist())
                tipos_selecionados = c_ex1.multiselect("🏷️ Filtrar TIPO DE NOTA (Opcional):", tipos_nota_unicos, default=tipos_nota_unicos)
                if not tipos_selecionados:
                    st.warning("Selecione pelo menos um Tipo de Nota para prosseguir."); return
                df_tasks_alocadas = df_tasks_alocadas[df_tasks_alocadas['TIPO NOTA'].astype(str).isin(tipos_selecionados)]

            todas_cols = df_tasks_alocadas.columns.tolist()
            cols_padrao = [c for c in ['PROTOCOLO', 'NOME DO SOLICITANTE', 'MUNICIPIO', 'TIPO LIGACAO', 'STATUS SAP', 'STATUS LIST', 'TIPO NOTA'] if c in todas_cols]
            colunas_exibir = c_ex1.multiselect("Colunas para aparecer no Balão do KML", todas_cols, default=cols_padrao)
            
            c_ex2.info("⚡ **Prioridade Automática Ativada:** Obras com TIPO NOTA igual a **CCF, DIF, MGD, MTP, ASC** ou **SID** recebem pino vermelho e são roteirizadas apenas para Equipes Principais.")
            col_prioridade = "TIPO NOTA"

    # === INÍCIO DO PROCESSAMENTO BIFÁSICO (TSP + OSRM) ===
    if st.button("🚀 Iniciar Motor de Roteirização (Processo em Nuvem)", type="primary", use_container_width=True):
        if df_tasks_alocadas.empty:
            st.error("Selecione equipes e regras compatíveis com a planilha primeiro."); return

        progresso_texto = st.empty()
        barra_progresso = st.progress(0)
        tempo_restante_texto = st.empty()
        
        start_time = time.time()
        api_calls = 0
        total_obras_rotear = len(df_tasks_alocadas)
        obras_processadas = 0
        obras_sobra_total = 0

        routed_data = []
        df_todas_bases_ativas = pd.DataFrame(bases_records)
        levantadores_unicos = list(set([b['LEVANTADOR'] for b in bases_records]))

        for b_name in levantadores_unicos:
            base_ref = df_todas_bases_ativas[df_todas_bases_ativas['LEVANTADOR'] == b_name].iloc[0]
            if pd.isna(base_ref.get('LATITUDE')): continue
            
            base_lat, base_lon = float(base_ref['LATITUDE']), float(base_ref['LONGITUDE'])
            unvisited = df_tasks_alocadas[df_tasks_alocadas['BASE_ATRIBUIDA'] == b_name].copy()
            
            periodo_atual = 1
            ordem_absoluta = 1
            
            while not unvisited.empty:
                dia_obras_prio = []
                dia_obras_norm = []
                tempo_dia = 0.0
                qtd_dia = 0
                curr_lat, curr_lon = base_lat, base_lon
                
                while not unvisited.empty:
                    unvisited_prio = unvisited[unvisited['PRIORIDADE'] == 'Sim']
                    if not unvisited_prio.empty:
                        dists = haversine_vectorized(curr_lat, curr_lon, unvisited_prio['LATITUDE'].values, unvisited_prio['LONGITUDE'].values)
                        nearest_idx = unvisited_prio.index[dists.argmin()]
                        is_prio = True
                    else:
                        dists = haversine_vectorized(curr_lat, curr_lon, unvisited['LATITUDE'].values, unvisited['LONGITUDE'].values)
                        nearest_idx = unvisited.index[dists.argmin()]
                        is_prio = False
                    
                    nearest_row = unvisited.loc[nearest_idx]
                    dist_km = round(dists.min(), 2)
                    
                    is_rural = False
                    if 'LOCALIDADE' in nearest_row and str(nearest_row['LOCALIDADE']).upper() == 'RURAL': is_rural = True
                    if 'TIPO NOTA' in nearest_row and str(nearest_row['TIPO NOTA']).upper() == 'UNR': is_rural = True
                    
                    tempo_viagem_h = (dist_km / velocidade_media_kmh) * (1.6 if is_rural else 1.0)
                    tempo_necessario = tempo_viagem_h + tempo_medio_obra
                    
                    if modo_limite == "Quantidade Fixa de Obras" and qtd_dia >= obras_por_periodo: break
                    if modo_limite != "Quantidade Fixa de Obras" and tempo_dia + tempo_necessario > horas_por_dia and qtd_dia > 0: break
                        
                    if is_prio: dia_obras_prio.append(nearest_row.to_dict())
                    else: dia_obras_norm.append(nearest_row.to_dict())
                    
                    curr_lat, curr_lon = nearest_row['LATITUDE'], nearest_row['LONGITUDE']
                    unvisited = unvisited.drop(nearest_idx)
                    tempo_dia += tempo_necessario
                    qtd_dia += 1
                    
                if len(dia_obras_prio) == 0 and len(dia_obras_norm) == 0: break
                
                progresso_texto.text(f"🧠 Otimizando sequência do {tipo_periodo} {periodo_atual} para {b_name} (TSP 2-Opt)...")
                
                last_prio_lat, last_prio_lon = base_lat, base_lon
                if len(dia_obras_prio) > 0:
                    last_prio_lat, last_prio_lon = dia_obras_prio[-1]['LATITUDE'], dia_obras_prio[-1]['LONGITUDE']
                    
                dia_obras_norm = otimizar_rota_tsp_2opt(dia_obras_norm, last_prio_lat, last_prio_lon)
                dia_final = dia_obras_prio + dia_obras_norm
                
                start_lat, start_lon = base_lat, base_lon
                almoco_inserido = False
                tempo_acumulado_rota = 0.0
                
                for obra in dia_final:
                    if modo_limite != "Quantidade Fixa de Obras" and tempo_acumulado_rota >= 4.0 and not almoco_inserido:
                        routed_data.append({
                            'PROTOCOLO': 'PAUSA_ALMOCO', 'NOME DO SOLICITANTE': '🍔 HORÁRIO DE ALMOÇO (1h)',
                            'LATITUDE': start_lat, 'LONGITUDE': start_lon, 'BASE_ATRIBUIDA': b_name, 'ORDEM': ordem_absoluta,
                            'SEMANA': periodo_atual if tipo_periodo == "Semana" else 1, 'DIA': periodo_atual if tipo_periodo == "Dia" else 1,
                            'PERIODO': periodo_atual, 'DISTANCIA_PONTO_ANTERIOR_KM': 0.0, 'TEMPO_VIAGEM_MINUTOS': 60.0,
                            'ROTA_GEOMETRIA': [[start_lon, start_lat], [start_lon, start_lat]], 'PRIORIDADE': 'Não'
                        })
                        almoco_inserido = True
                        ordem_absoluta += 1
                        tempo_acumulado_rota += 1.0
                        
                    progresso_texto.text(f"🗺️ Mapeando {b_name} | {tipo_periodo} {periodo_atual} | Obra {ordem_absoluta} via Satélite...")
                    rota_geom, dur_sec = obter_rota_ruas(start_lat, start_lon, obra['LATITUDE'], obra['LONGITUDE'], velocidade_media_kmh)
                    api_calls += 1
                    
                    is_rur = False
                    if 'LOCALIDADE' in obra and str(obra['LOCALIDADE']).upper() == 'RURAL': is_rur = True
                    if 'TIPO NOTA' in obra and str(obra['TIPO NOTA']).upper() == 'UNR': is_rur = True
                    if is_rur: dur_sec *= 1.6
                    
                    obra['ORDEM'] = ordem_absoluta
                    obra['SEMANA'] = periodo_atual if tipo_periodo == "Semana" else 1
                    obra['DIA'] = periodo_atual if tipo_periodo == "Dia" else 1
                    obra['PERIODO'] = periodo_atual
                    obra['DISTANCIA_PONTO_ANTERIOR_KM'] = round(haversine_vectorized(start_lat, start_lon, obra['LATITUDE'], obra['LONGITUDE']), 2)
                    obra['TEMPO_VIAGEM_MINUTOS'] = round(dur_sec / 60.0, 1)
                    obra['ROTA_GEOMETRIA'] = rota_geom
                    
                    routed_data.append(obra)
                    start_lat, start_lon = obra['LATITUDE'], obra['LONGITUDE']
                    ordem_absoluta += 1
                    tempo_acumulado_rota += (dur_sec / 3600.0) + tempo_medio_obra
                    obras_processadas += 1
                    
                    barra_progresso.progress(min(obras_processadas / total_obras_rotear, 1.0))
                    
                    elapsed = time.time() - start_time
                    avg_time = elapsed / api_calls
                    obras_restantes = total_obras_rotear - obras_processadas
                    if obras_restantes > 0:
                        est_rem = avg_time * obras_restantes
                        m, s = divmod(int(est_rem), 60)
                        h, m = divmod(m, 60)
                        if h > 0: tempo_restante_texto.markdown(f"⏳ **Tempo estimado restante:** {h:02d}h {m:02d}m {s:02d}s")
                        else: tempo_restante_texto.markdown(f"⏳ **Tempo estimado restante:** {m:02d}m {s:02d}s")
                    else: tempo_restante_texto.markdown("✅ **Processamento Concluído! Montando arquivos...**")
                    time.sleep(1.2)
                    
                progresso_texto.text(f"🏠 Encerrando pacote de {b_name}, traçando retorno final...")
                rota_retorno, dur_ret_seg = obter_rota_ruas(start_lat, start_lon, base_lat, base_lon, velocidade_media_kmh)
                api_calls += 1
                dist_retorno = haversine_vectorized(start_lat, start_lon, base_lat, base_lon)
                routed_data.append({
                    'PROTOCOLO': 'RETORNO_BASE', 'NOME DO SOLICITANTE': 'BASE_RETORNO', 'LATITUDE': base_lat, 'LONGITUDE': base_lon,
                    'BASE_ATRIBUIDA': b_name, 'ORDEM': ordem_absoluta, 'SEMANA': periodo_atual if tipo_periodo == "Semana" else 1,
                    'DIA': periodo_atual if tipo_periodo == "Dia" else 1, 'PERIODO': periodo_atual,
                    'DISTANCIA_PONTO_ANTERIOR_KM': round(dist_retorno, 2), 'TEMPO_VIAGEM_MINUTOS': round(dur_ret_seg / 60.0, 1),
                    'ROTA_GEOMETRIA': rota_retorno, 'PRIORIDADE': 'Não'
                })
                time.sleep(1.2)
                
                periodo_atual += 1
                ordem_absoluta = 1
                if periodo_atual > limite_periodos:
                    obras_sobra_total += len(unvisited)
                    obras_processadas += len(unvisited) 
                    break

        if obras_sobra_total > 0:
            st.warning(f"⏳ {obras_sobra_total} obras ficaram de fora do roteiro porque a carga horária/limite estourou.")

        df_final_route = pd.DataFrame(routed_data)
        df_final_route['DISTANCIA_PROXIMO_PONTO_KM'] = df_final_route.groupby(['BASE_ATRIBUIDA', 'PERIODO'])['DISTANCIA_PONTO_ANTERIOR_KM'].shift(-1).fillna(0.0)
        
        st.session_state.df_routed = df_final_route
        st.session_state.bases_records = bases_records
        st.session_state.tipo_periodo = tipo_periodo
        st.session_state.colunas_exibir = colunas_exibir
        st.session_state.col_prioridade = col_prioridade
        st.session_state.roteamento_concluido = True
        
        st.rerun()
